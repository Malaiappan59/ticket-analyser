# ─────────────────────────────────────────────────────────────────────────────
# core/exporter.py  –  Generate the classified, multi-sheet Excel output
# ─────────────────────────────────────────────────────────────────────────────

from __future__ import annotations

from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from config.settings import CATEGORIES, CATEGORY_COLORS

# ─────────────────────────────────────────────────────────────────────────────
# Shared style helpers
# ─────────────────────────────────────────────────────────────────────────────

_NAVY       = "0D47A1"
_WHITE      = "FFFFFF"
_ALT_ROW    = "F0F4FF"
_TOTAL_BG   = "E8EAF6"
_TOTAL_FG   = "0D47A1"
_NOTE_FG    = "757575"

_MEDIUM_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _header_cell(
    ws,
    row: int,
    col: int,
    value: str,
    bg: str = _NAVY,
    fg: str = _WHITE,
    size: int = 10,
    bold: bool = True,
    wrap: bool = True,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(bg)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = _MEDIUM_BORDER


def _data_cell(
    ws,
    row: int,
    col: int,
    value,
    alt_row: bool = False,
    bold: bool = False,
    wrap: bool = False,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if alt_row:
        cell.fill = _fill(_ALT_ROW)
    cell.font = Font(bold=bold, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = _THIN_BORDER


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def create_output_excel(
    original_df: pd.DataFrame,
    classified_df: pd.DataFrame,
    category_col: str = "Category",
) -> BytesIO:
    """
    Build a formatted 3-sheet Excel workbook and return it as a BytesIO object.

    Sheet 1 – "Classified Tickets"
        Columns = IT categories.  Each column lists the ticket IDs belonging
        to that category.  A "Total: N" row closes each column.

    Sheet 2 – "All Tickets (Classified)"
        Full original data with a highlighted "Category" column prepended.
        Auto-filter enabled.

    Sheet 3 – "Category Summary"
        Count + percentage table per category plus a bar chart.
    """
    wb = Workbook()
    # Remove default empty sheet
    for sheet in wb.worksheets:
        wb.remove(sheet)

    _build_classified_sheet(wb, classified_df, category_col)
    _build_all_tickets_sheet(wb, classified_df)
    _build_summary_sheet(wb, classified_df, category_col)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_classified_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    category_col: str,
) -> None:
    """
    Sheet 1: category-column layout.
    Ticket IDs are listed under their category column header.
    """
    ws = wb.create_sheet("Classified Tickets")

    # Prepare data: category → list of ticket IDs (first column or index)
    id_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]  # skip Category col
    # If "Category" is col 0, ticket ID is col 1
    if df.columns[0] == category_col and len(df.columns) > 1:
        id_col = df.columns[1]
    else:
        id_col = df.columns[0]

    cats = list(CATEGORIES.keys())
    cat_tickets: dict[str, list[str]] = {c: [] for c in cats}

    for _, row in df.iterrows():
        cat = str(row.get(category_col, "Others"))
        cat = cat if cat in cat_tickets else "Others"
        cat_tickets[cat].append(str(row.get(id_col, "")))

    max_rows = max((len(v) for v in cat_tickets.values()), default=0)

    # ── Column headers ──────────────────────────────────────────────────────
    for c_idx, cat in enumerate(cats, 1):
        color = CATEGORY_COLORS.get(cat, "CCCCCC")
        _header_cell(ws, 1, c_idx, cat, bg=color, fg="1A1A1A", size=11)
        ws.column_dimensions[get_column_letter(c_idx)].width = 22

    ws.row_dimensions[1].height = 28

    # ── Data rows ───────────────────────────────────────────────────────────
    for c_idx, cat in enumerate(cats, 1):
        for r_idx, tid in enumerate(cat_tickets[cat], 2):
            _data_cell(ws, r_idx, c_idx, tid, alt_row=(r_idx % 2 == 0))

    # ── Totals row ──────────────────────────────────────────────────────────
    total_row = max_rows + 2
    for c_idx, cat in enumerate(cats, 1):
        count = len(cat_tickets[cat])
        cell = ws.cell(row=total_row, column=c_idx, value=f"Total: {count}")
        cell.fill = _fill(_TOTAL_BG)
        cell.font = Font(bold=True, size=11, color=_TOTAL_FG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _MEDIUM_BORDER

    ws.row_dimensions[total_row].height = 20

    # ── Footer note ─────────────────────────────────────────────────────────
    note_row = total_row + 2
    note_cell = ws.cell(
        row=note_row,
        column=1,
        value=(
            "ℹ  Ticket IDs are listed under each category column.  "
            "Totals shown in the last data row.  "
            "'Others' = unclassified tickets."
        ),
    )
    note_cell.font = Font(italic=True, size=9, color=_NOTE_FG)
    ws.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row, end_column=min(len(cats), 12),
    )
    ws.freeze_panes = "A2"


def _build_all_tickets_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """
    Sheet 2: all original columns + Category column, with auto-filter.
    Category cell background matches the category colour.
    """
    ws = wb.create_sheet("All Tickets (Classified)")

    cols = list(df.columns)

    # ── Header row ──────────────────────────────────────────────────────────
    for c_idx, col in enumerate(cols, 1):
        label = col.replace("_", " ").title()
        _header_cell(ws, 1, c_idx, label, size=10)
        # Auto-width estimate
        max_w = max(len(label), 12)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_w + 4, 40)

    ws.row_dimensions[1].height = 24

    # ── Data rows ───────────────────────────────────────────────────────────
    cat_col_idx = (cols.index("Category") + 1) if "Category" in cols else None

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        alt = r_idx % 2 == 0
        for c_idx, col in enumerate(cols, 1):
            val = row[col]
            val = "" if pd.isna(val) or str(val) == "nan" else str(val)

            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=False
            )
            cell.font = Font(size=9)
            if alt:
                cell.fill = _fill(_ALT_ROW)
            cell.border = _THIN_BORDER

            # Colour the Category cell
            if c_idx == cat_col_idx:
                color = CATEGORY_COLORS.get(val, "EEEEEE")
                cell.fill = _fill(color)
                cell.font = Font(size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"
    ws.freeze_panes = "A2"


def _build_summary_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    category_col: str,
) -> None:
    """
    Sheet 3: per-category count + percentage table + bar chart.
    """
    ws = wb.create_sheet("Category Summary")

    cats     = list(CATEGORIES.keys())
    total    = len(df)
    counts   = df[category_col].value_counts() if category_col in df.columns else {}

    # ── Table header ────────────────────────────────────────────────────────
    headers = ["#", "Category", "Ticket Count", "Percentage", "Visual Bar"]
    col_widths = [5, 20, 14, 13, 30]
    for c_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        _header_cell(ws, 1, c_idx, h, size=10)
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.row_dimensions[1].height = 26

    # ── Data rows ───────────────────────────────────────────────────────────
    for r_idx, cat in enumerate(cats, 2):
        count = int(counts.get(cat, 0))
        pct   = (count / total * 100) if total else 0
        bar   = "█" * int(pct / 2)       # rough visual bar

        alt = r_idx % 2 == 0

        # Serial number
        _data_cell(ws, r_idx, 1, r_idx - 1, alt_row=alt)

        # Category (coloured background)
        cat_cell = ws.cell(row=r_idx, column=2, value=cat)
        cat_cell.fill = _fill(CATEGORY_COLORS.get(cat, "EEEEEE"))
        cat_cell.font = Font(bold=True, size=10)
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        cat_cell.border = _THIN_BORDER

        # Count
        cnt_cell = ws.cell(row=r_idx, column=3, value=count)
        cnt_cell.font = Font(size=10, bold=(count == max((int(counts.get(c, 0)) for c in cats), default=0)))
        cnt_cell.alignment = Alignment(horizontal="center", vertical="center")
        cnt_cell.border = _THIN_BORDER
        if alt:
            cnt_cell.fill = _fill(_ALT_ROW)

        # Percentage
        pct_cell = ws.cell(row=r_idx, column=4, value=f"{pct:.1f}%")
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
        pct_cell.border = _THIN_BORDER
        pct_cell.font = Font(size=10)
        if alt:
            pct_cell.fill = _fill(_ALT_ROW)

        # Bar
        bar_cell = ws.cell(row=r_idx, column=5, value=bar)
        bar_cell.font = Font(size=9, color="0D47A1")
        bar_cell.alignment = Alignment(horizontal="left", vertical="center")
        bar_cell.border = _THIN_BORDER
        if alt:
            bar_cell.fill = _fill(_ALT_ROW)

    # ── TOTAL row ────────────────────────────────────────────────────────────
    total_row = len(cats) + 2
    for c_idx, val in enumerate(["", "TOTAL", total, "100.0%", ""], 1):
        cell = ws.cell(row=total_row, column=c_idx, value=val)
        cell.fill = _fill(_NAVY)
        cell.font = Font(bold=True, color=_WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _MEDIUM_BORDER

    ws.row_dimensions[total_row].height = 24

    # ── Bar chart ────────────────────────────────────────────────────────────
    chart = BarChart()
    chart.type   = "col"
    chart.title  = "Ticket Volume by Category"
    chart.y_axis.title = "Ticket Count"
    chart.x_axis.title = "Category"
    chart.style  = 10
    chart.width  = 22
    chart.height = 14

    data_ref = Reference(ws, min_col=3, min_row=1, max_row=len(cats) + 1)
    cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(cats) + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape  = 4

    ws.add_chart(chart, "G2")

    ws.freeze_panes = "A2"
