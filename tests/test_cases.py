"""
tests/test_cases.py
--------------------
Feasibility validation with best/worst edge cases.
Run:  python -m pytest tests/test_cases.py -v
      or
      python tests/test_cases.py  (no pytest required)
"""

from __future__ import annotations

import io
import sys
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent))

from config.settings import CATEGORIES
from core.classifier import classify_with_keywords, _normalise_category
from core.preprocessor import (
    apply_filters,
    clean_dataframe,
    detect_columns,
    get_filter_options,
    load_file,
    validate_dataframe,
)
from core.exporter import create_output_excel
from utils.helpers import generate_sample_data


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _make_df(**cols) -> pd.DataFrame:
    """Quick DataFrame factory."""
    return pd.DataFrame(cols)


def _csv_file(content: str, name: str = "test.csv"):
    """Simulate a Streamlit UploadedFile for CSV."""
    buf = io.BytesIO(content.encode("utf-8"))
    buf.name = name
    buf.read = buf.read  # already has .read()
    # Patch .name attribute
    buf.name = name
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# 1. Keyword Classifier Tests
# ─────────────────────────────────────────────────────────────────────────────

class TestKeywordClassifier(unittest.TestCase):

    # ── BEST cases (clear, unambiguous) ──────────────────────────────────────
    def test_cpu_clear(self):
        cat = classify_with_keywords("High CPU usage", "Server CPU at 98% for 30 minutes")
        self.assertEqual(cat, "CPU")

    def test_memory_clear(self):
        cat = classify_with_keywords("Out of memory error", "JVM heap exhausted, OOM killer triggered")
        self.assertEqual(cat, "Memory")

    def test_storage_clear(self):
        cat = classify_with_keywords("Disk space critical", "Disk usage at 96% on /var partition")
        self.assertEqual(cat, "Storage")

    def test_network_clear(self):
        cat = classify_with_keywords("Network connectivity failure", "VPN tunnel down, DNS not resolving")
        self.assertEqual(cat, "Network")

    def test_database_clear(self):
        cat = classify_with_keywords("Oracle tablespace full", "APP_DATA tablespace at 97% capacity")
        self.assertEqual(cat, "Database")

    def test_security_clear(self):
        cat = classify_with_keywords("SSL certificate expired", "TLS cert on api.company.com expired")
        self.assertEqual(cat, "Security")

    def test_hardware_clear(self):
        cat = classify_with_keywords("Server fan failure", "Physical server thermal alert, hardware fault")
        self.assertEqual(cat, "Hardware")

    def test_os_clear(self):
        cat = classify_with_keywords("Linux kernel panic", "Server crashed with kernel panic message")
        self.assertEqual(cat, "OS")

    def test_monitoring_clear(self):
        cat = classify_with_keywords("Datadog alert fired", "Splunk monitoring alert threshold breached")
        self.assertEqual(cat, "Monitoring")

    def test_middleware_clear(self):
        cat = classify_with_keywords("Tomcat not responding", "Apache Tomcat thread pool exhausted")
        self.assertEqual(cat, "Middleware")

    def test_application_clear(self):
        cat = classify_with_keywords("Application crash", "Service process crashed with unhandled exception")
        self.assertEqual(cat, "Application")

    # ── WORST/Edge cases ─────────────────────────────────────────────────────
    def test_empty_strings_returns_others(self):
        """Completely empty ticket → Others."""
        cat = classify_with_keywords("", "")
        self.assertEqual(cat, "Others")

    def test_whitespace_only_returns_others(self):
        cat = classify_with_keywords("   ", "\t\n  ")
        self.assertEqual(cat, "Others")

    def test_none_values_returns_others(self):
        cat = classify_with_keywords(None, None)  # type: ignore[arg-type]
        self.assertEqual(cat, "Others")

    def test_gibberish_returns_others(self):
        cat = classify_with_keywords("xkqzw mfplv", "abc123 zzzzz")
        self.assertEqual(cat, "Others")

    def test_special_characters_only(self):
        cat = classify_with_keywords("!@#$%^&*()", ">>>><<<[]{}|\\")
        self.assertEqual(cat, "Others")

    def test_very_long_description(self):
        """1500-char description should still classify correctly."""
        long_desc = "cpu processor high utilization " * 50
        cat = classify_with_keywords("CPU issue", long_desc)
        self.assertEqual(cat, "CPU")

    def test_mixed_language_with_keywords(self):
        """Ticket with some non-English text but English keywords."""
        cat = classify_with_keywords("Memory issue détectée", "ram usage sehr hoch 95% memory")
        self.assertEqual(cat, "Memory")

    def test_ambiguous_ticket_returns_valid_category(self):
        """Ambiguous description – must return SOME valid category, not crash."""
        cat = classify_with_keywords("Server issue", "The server is not working properly")
        self.assertIn(cat, list(CATEGORIES.keys()))

    def test_multiple_category_keywords_highest_wins(self):
        """CPU gets more hits → wins over Memory."""
        cat = classify_with_keywords(
            "CPU memory issue",
            "cpu cpu cpu cpu memory memory"
        )
        self.assertEqual(cat, "CPU")

    def test_unicode_emoji_in_description(self):
        """Emoji-laden ticket shouldn't crash."""
        cat = classify_with_keywords("🔴 CPU 🔥 HIGH", "💥 processor overload 🚨")
        self.assertIn(cat, list(CATEGORIES.keys()))

    def test_numeric_only_description(self):
        cat = classify_with_keywords("12345", "9876543210")
        self.assertEqual(cat, "Others")


# ─────────────────────────────────────────────────────────────────────────────
# 2. Category Normalisation Tests
# ─────────────────────────────────────────────────────────────────────────────

class TestNormaliseCategory(unittest.TestCase):

    def test_exact_match(self):
        self.assertEqual(_normalise_category("CPU"), "CPU")
        self.assertEqual(_normalise_category("Memory"), "Memory")

    def test_case_insensitive(self):
        self.assertEqual(_normalise_category("cpu"), "CPU")
        self.assertEqual(_normalise_category("MEMORY"), "Memory")

    def test_with_newline(self):
        self.assertEqual(_normalise_category("Storage\n"), "Storage")

    def test_with_punctuation(self):
        self.assertEqual(_normalise_category("Network."), "Network")

    def test_unknown_returns_others(self):
        self.assertEqual(_normalise_category("Plumbing"), "Others")

    def test_empty_returns_others(self):
        self.assertEqual(_normalise_category(""), "Others")

    def test_partial_match(self):
        # "Databases" → Database via substring
        result = _normalise_category("Databases")
        self.assertIn(result, list(CATEGORIES.keys()))


# ─────────────────────────────────────────────────────────────────────────────
# 3. Preprocessor – File Loading
# ─────────────────────────────────────────────────────────────────────────────

class TestPreprocessorLoading(unittest.TestCase):

    def _make_uploaded(self, content: bytes, name: str):
        buf = io.BytesIO(content)
        buf.name = name
        return buf

    def test_valid_csv(self):
        csv = b"number,short_description,status\nINC001,CPU high,New\nINC002,Disk full,Closed"
        f   = self._make_uploaded(csv, "tickets.csv")
        df, msg = load_file(f)
        self.assertIsNotNone(df)
        self.assertEqual(len(df), 2)

    def test_unsupported_extension(self):
        f = self._make_uploaded(b"irrelevant", "tickets.pdf")
        df, msg = load_file(f)
        self.assertIsNone(df)
        self.assertIn("Unsupported", msg)

    def test_empty_csv(self):
        f = self._make_uploaded(b"col1,col2\n", "empty.csv")
        df, msg = load_file(f)
        self.assertIsNone(df)

    def test_column_normalisation(self):
        csv = b"Short Description,Assignment Group,Ticket Number\nCPU,L1 Team,INC001"
        f   = self._make_uploaded(csv, "t.csv")
        df, _ = load_file(f)
        self.assertIsNotNone(df)
        # Columns should be snake_case
        self.assertIn("short_description", df.columns)
        self.assertIn("assignment_group", df.columns)

    def test_csv_with_bom(self):
        """UTF-8 BOM should be handled."""
        csv = "\ufeffnumber,description\nINC001,CPU high".encode("utf-8-sig")
        f   = self._make_uploaded(csv, "bom.csv")
        df, msg = load_file(f)
        self.assertIsNotNone(df)

    def test_latin1_encoding(self):
        """Latin-1 encoded CSV should still load."""
        csv = "number,description\nINC001,Disqu\xe9 plein".encode("latin-1")
        f   = self._make_uploaded(csv, "latin.csv")
        df, msg = load_file(f)
        self.assertIsNotNone(df)


# ─────────────────────────────────────────────────────────────────────────────
# 4. Preprocessor – Column Detection
# ─────────────────────────────────────────────────────────────────────────────

class TestColumnDetection(unittest.TestCase):

    def test_standard_servicenow_columns(self):
        df = _make_df(
            number=["INC001"],
            short_description=["CPU high"],
            status=["New"],
            assignment_group=["L1"],
        )
        detected = detect_columns(df)
        self.assertEqual(detected.get("id"), "number")
        self.assertEqual(detected.get("short_description"), "short_description")
        self.assertEqual(detected.get("status"), "status")
        self.assertEqual(detected.get("assignment_group"), "assignment_group")

    def test_no_matching_columns(self):
        df = _make_df(col_a=["x"], col_b=["y"])
        detected = detect_columns(df)
        self.assertEqual(detected, {})

    def test_partial_detection(self):
        df = _make_df(status=["open"], foo=["bar"])
        detected = detect_columns(df)
        self.assertIn("status", detected)


# ─────────────────────────────────────────────────────────────────────────────
# 5. Preprocessor – Filtering
# ─────────────────────────────────────────────────────────────────────────────

class TestFiltering(unittest.TestCase):

    def setUp(self):
        self.df = pd.DataFrame({
            "Category":         ["CPU", "Memory", "Storage", "CPU", "Others"],
            "number":           ["INC001", "INC002", "INC003", "INC004", "INC005"],
            "status":           ["New", "Closed", "In Progress", "Resolved", "New"],
            "assignment_group": ["L1", "L2-Server", "L2-Network", "L1", "L2-DB"],
        })

    def test_filter_by_status(self):
        result = apply_filters(self.df, status="New", status_col="status")
        self.assertEqual(len(result), 2)

    def test_filter_by_category(self):
        result = apply_filters(self.df, category=["CPU"], category_col="Category")
        self.assertEqual(len(result), 2)

    def test_filter_by_group(self):
        result = apply_filters(
            self.df, assignment_group="L1", assignment_group_col="assignment_group"
        )
        self.assertEqual(len(result), 2)

    def test_filter_by_ticket_id(self):
        result = apply_filters(self.df, ticket_id="INC003", id_col="number")
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]["number"], "INC003")

    def test_no_filter_returns_all(self):
        result = apply_filters(self.df)
        self.assertEqual(len(result), 5)

    def test_combined_filters(self):
        result = apply_filters(
            self.df,
            status="New",
            category=["CPU"],
            status_col="status",
            category_col="Category",
        )
        self.assertEqual(len(result), 1)

    def test_filter_nonexistent_group(self):
        result = apply_filters(
            self.df, assignment_group="ZZZZ", assignment_group_col="assignment_group"
        )
        self.assertEqual(len(result), 0)


# ─────────────────────────────────────────────────────────────────────────────
# 6. Validation
# ─────────────────────────────────────────────────────────────────────────────

class TestValidation(unittest.TestCase):

    def test_valid_df(self):
        df = pd.DataFrame({"a": range(10), "b": range(10)})
        ok, issues = validate_dataframe(df)
        self.assertTrue(ok)

    def test_empty_df(self):
        df = pd.DataFrame()
        ok, issues = validate_dataframe(df)
        self.assertFalse(ok)
        self.assertTrue(any("❌" in i for i in issues))

    def test_single_column_df(self):
        df = pd.DataFrame({"only_col": [1, 2, 3]})
        ok, issues = validate_dataframe(df)
        self.assertFalse(ok)

    def test_large_df_warning(self):
        df = pd.DataFrame({"a": range(6000), "b": range(6000)})
        ok, issues = validate_dataframe(df)
        self.assertTrue(ok)  # no error
        self.assertTrue(any("⚠️" in i for i in issues))

    def test_duplicate_rows_warning(self):
        df = pd.DataFrame({"a": [1, 1, 2], "b": ["x", "x", "y"]})
        ok, issues = validate_dataframe(df)
        self.assertTrue(any("duplicate" in i.lower() for i in issues))


# ─────────────────────────────────────────────────────────────────────────────
# 7. Excel Exporter
# ─────────────────────────────────────────────────────────────────────────────

class TestExporter(unittest.TestCase):

    def _make_classified_df(self, n: int = 50) -> pd.DataFrame:
        sample = generate_sample_data(n)
        # Simulate running keyword classifier on it
        from core.classifier import classify_with_keywords
        cats = [
            classify_with_keywords(
                row.get("Short_Description", ""),
                row.get("Description", ""),
            )
            for _, row in sample.iterrows()
        ]
        sample.insert(0, "Category", cats)
        return sample

    def test_excel_is_generated(self):
        df = self._make_classified_df(30)
        orig = df.drop(columns=["Category"])
        output = create_output_excel(orig, df, "Category")
        self.assertIsNotNone(output)
        self.assertGreater(output.getbuffer().nbytes, 0)

    def test_excel_has_three_sheets(self):
        from openpyxl import load_workbook
        df   = self._make_classified_df(20)
        orig = df.drop(columns=["Category"])
        out  = create_output_excel(orig, df, "Category")
        wb   = load_workbook(out)
        self.assertEqual(len(wb.sheetnames), 3)
        self.assertIn("Classified Tickets",     wb.sheetnames)
        self.assertIn("All Tickets (Classified)", wb.sheetnames)
        self.assertIn("Category Summary",       wb.sheetnames)

    def test_excel_large_dataset(self):
        """Excel generation should not raise for 700+ rows."""
        df = self._make_classified_df(750)
        orig = df.drop(columns=["Category"])
        try:
            out = create_output_excel(orig, df, "Category")
            self.assertGreater(out.getbuffer().nbytes, 0)
        except Exception as exc:
            self.fail(f"Excel generation raised: {exc}")

    def test_excel_with_all_others(self):
        """Edge case: all tickets classified as Others."""
        df = pd.DataFrame({
            "Category":    ["Others"] * 10,
            "Number":      [f"INC{i:04d}" for i in range(10)],
            "Description": ["unknown"] * 10,
        })
        orig = df.drop(columns=["Category"])
        out  = create_output_excel(orig, df, "Category")
        self.assertGreater(out.getbuffer().nbytes, 0)

    def test_excel_with_empty_classified_df(self):
        """Edge case: zero rows."""
        df = pd.DataFrame({"Category": [], "Number": [], "Description": []})
        orig = df.drop(columns=["Category"])
        try:
            create_output_excel(orig, df, "Category")
        except Exception as exc:
            self.fail(f"Empty DF raised: {exc}")


# ─────────────────────────────────────────────────────────────────────────────
# 8. Sample Data Generator
# ─────────────────────────────────────────────────────────────────────────────

class TestSampleDataGenerator(unittest.TestCase):

    def test_generates_correct_number(self):
        df = generate_sample_data(100)
        self.assertEqual(len(df), 100)

    def test_generates_default_columns(self):
        df = generate_sample_data(10)
        for col in ["Number", "Type", "Short_Description", "Description", "Status"]:
            self.assertIn(col, df.columns)

    def test_no_null_required_fields(self):
        df = generate_sample_data(50)
        for col in ["Number", "Short_Description", "Status"]:
            self.assertFalse(df[col].isnull().any(), f"Null in {col}")

    def test_unique_ticket_numbers(self):
        df = generate_sample_data(200)
        self.assertEqual(df["Number"].nunique(), 200)

    def test_status_valid_values(self):
        valid = {"New", "In Progress", "On Hold", "Resolved", "Closed"}
        df    = generate_sample_data(100)
        self.assertTrue(set(df["Status"].unique()).issubset(valid))

    def test_750_tickets_performance(self):
        """Generating 750 tickets should complete quickly."""
        import time
        t0 = time.time()
        generate_sample_data(750)
        elapsed = time.time() - t0
        self.assertLess(elapsed, 5.0, "Sample generation took > 5 seconds!")


# ─────────────────────────────────────────────────────────────────────────────
# 9. End-to-End Pipeline (Keyword mode only – no Ollama needed)
# ─────────────────────────────────────────────────────────────────────────────

class TestEndToEndPipeline(unittest.TestCase):

    def test_full_pipeline_keyword_mode(self):
        """Upload → detect columns → classify → export without errors."""
        from core.classifier import classify_batch

        # Generate sample
        df = generate_sample_data(50)
        df.columns = df.columns.str.lower().str.replace(" ", "_")
        df = clean_dataframe(df)

        # Detect columns
        detected = detect_columns(df)

        # Classify
        cats = classify_batch(
            tickets_df=df,
            short_desc_col=detected.get("short_description"),
            desc_col=detected.get("description"),
            type_col=detected.get("type"),
            use_llm=False,
        )
        self.assertEqual(len(cats), len(df))
        for c in cats:
            self.assertIn(c, list(CATEGORIES.keys()))

        # Build Excel
        classified = df.copy()
        classified.insert(0, "Category", cats)
        out = create_output_excel(df, classified, "Category")
        self.assertGreater(out.getbuffer().nbytes, 0)

    def test_pipeline_with_missing_columns(self):
        """Pipeline should not crash when columns are missing."""
        from core.classifier import classify_batch

        df = pd.DataFrame({
            "ticket_id": ["T001", "T002", "T003"],
            "notes":     ["cpu usage high", "memory leak", "network down"],
        })
        df = clean_dataframe(df)

        cats = classify_batch(
            tickets_df=df,
            short_desc_col=None,
            desc_col="notes",
            type_col=None,
            use_llm=False,
        )
        self.assertEqual(len(cats), 3)

    def test_pipeline_with_all_empty_descriptions(self):
        """All blank descriptions → all 'Others'."""
        from core.classifier import classify_batch

        df = pd.DataFrame({
            "id":   range(5),
            "desc": ["", "  ", "nan", "", ""],
        })
        df = clean_dataframe(df)
        cats = classify_batch(
            tickets_df=df,
            short_desc_col=None,
            desc_col="desc",
            type_col=None,
            use_llm=False,
        )
        self.assertTrue(all(c == "Others" for c in cats))


# ─────────────────────────────────────────────────────────────────────────────
# Runner
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    loader = unittest.TestLoader()
    suite  = loader.loadTestsFromModule(sys.modules[__name__])
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    total  = result.testsRun
    passed = total - len(result.failures) - len(result.errors)
    print(f"\n{'='*60}")
    print(f"Results: {passed}/{total} passed | "
          f"{len(result.failures)} failed | {len(result.errors)} errors")
    sys.exit(0 if result.wasSuccessful() else 1)
